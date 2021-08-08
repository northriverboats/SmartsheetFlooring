#!/usr/bin/env python3

import click
import glob
import logging
import openpyxl
import os
import smartsheet
import subprocess
import sys
import time
from emailer import *
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

api = os.getenv('SMARTSHEET_API')
source_dir = os.getenv('SOURCE_DIR')
target_dir = os.getenv('TARGET_DIR')

reports = [
    {'id': 7487692708571012, 'name': 'Boat Country Flooring'},
    {'id': 2403688380688260, 'name': 'Clemens Flooring'},
    {'id': 5706277981579140, 'name': 'Idaho Marine Flooring'},
    {'id': 363956872210308,  'name': 'Port Boat House Flooring'},
    {'id': 7470375400433540, 'name': 'Valley Marine Flooring'},
    {'id': 8526731196819332, 'name': 'Y-Marina Flooring'},
]

log_text = ""
errors = False

def log(text, error=None):
    global log_text, errors
    print(text)
    log_text += text + "\n"
    if (error):
        errors = True

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def mail_results(subject, body):
    mFrom = os.getenv('MAIL_FROM')
    mTo = os.getenv('MAIL_TO')
    m = Email(os.getenv('MAIL_SERVER'))
    m.setFrom(mFrom)
    m.addRecipient(mTo)
    m.addCC(os.getenv('MAIL_ALSO'))

    m.setSubject(subject)
    m.setTextBody("You should not see this text in a MIME aware reader")
    m.setHtmlBody('<pre>\n' + body + '</pre>\n')
    m.send()


def fetch_value(cell):
    value = cell.value
    if cell.data_type == 's':
        return value
    if cell.is_date:
        return ('%02d/%02d/%02d' %(value.month,value.day,value.year-2000))
    if value == None:
        return ''
    return str(int(value))

def set_header(wsNew, row):
    titles = ['Hull #',
              'Primary',
              'Boat Model',
              'Order Details',
              'Flooring',
              'Invoice Amount',
              'Est Start/Finish',
              'Actual Start',
              'Actual Finish',
             ]
    wsNew.row_dimensions[row+7].height = 21.6
    for i in range(1,11):
        wsNew.cell(row=row+7, column=i,value=titles[i-1])
        wsNew.cell(row=row+7, column=i).alignment = Alignment(horizontal='center',vertical='center')

def process_row(wsOld,wsNew,row):
    for i in range(1,11):
        value = fetch_value(wsOld.cell(column=i,row=row))
        cell = wsNew.cell(column=i,row=row)
        cell.value = value

def process_rows(wsOld,wsNew):
    for i in range(2,wsOld.max_row+1):
        process_row(wsOld,wsNew,i)

    # set_footer(wsNew, wsOld.max_row+1)

def process_sheet_to_xlsx(file):
    # change variables here
    input_name = source_dir + 'downloads/' + file
    output_name = target_dir + file

    # load sheet data is coming from
    wbOld = openpyxl.load_workbook(input_name)
    wsOld = wbOld.active

    # load sheet we are copying data to
    wbNew = openpyxl.load_workbook(source_dir + 'FlooringTemplate.xlsx')
    wsNew = wbNew.active

    process_rows(wsOld, wsNew)

    range = 'A1:J'+str(wsOld.max_row)

    wbNew.create_named_range('_xlnm.Print_Area', wsNew, range, scope=0)

    # save new sheet out to new file
    try:
        wbNew.save(output_name)
    except Exception as e:
        log('             FAILED TO CREATE XLSX: ' + str(e), True)


def process_sheets():
    log("\nPROCESS SHEETS ===============================")
    os.chdir(source_dir + 'downloads/')
    for file in sorted(glob.glob('*.xlsx')):
        log("  converting %s" % (file))
        process_sheet_to_xlsx(file)


def download_sheets():
    files = os.listdir(source_dir + 'downloads')
    for file in files:
        os.remove(os.path.join(source_dir + 'downloads', file))

    smart = smartsheet.Smartsheet(api)
    smart.assume_user(os.getenv('SMARTSHEET_USER'))
    log("DOWNLOADING SHEETS ===========================")
    for report in reports:
        log("  downloading sheet: " + report['name'])
        try:
            smart.Reports.get_report_as_excel(report['id'], source_dir + 'downloads')
        except Exception as e:
            log('                     ERROR DOWNLOADING SHEET: ' + str(e), True)


def send_error_report():
    subject = 'Smartsheet Flooring Error Report'
    mail_results(subject, log_text)


@click.command()
@click.option(
    '--list',
    '-l',
    'list_',
    is_flag=True,
    help='Print list of dealers'
)
@click.option(
    '--dealer',
    '-d',
    multiple=True,
    help='Dealer to include (can use multiple times)'
)
@click.option(
    '--ignore',
    '-i',
    multiple=True,
    help='Dealer to ignore (can use multiple times)'
)
@click.option(
    '--download/--no-download',
    default=True,
    help='Download spreedsheets unless --no-download'
)
@click.option(
    '--excel/--no-excel',
    default=True,
    help='Create Excel Sheets unless --no-excel'
)
def main(list_, dealer, ignore, download, excel):
    # load environmental variables
    env_path = resource_path('.env')
    load_dotenv(dotenv_path=env_path)

    try:
        if download:
            download_sheets()
        if excel:
            process_sheets()
    except Exception as e:
        log('Uncaught Error in main(): ' + str(e), True)
    if (errors):
        send_error_report()
    sys.exit(0)


if __name__ == "__main__":
    main()
